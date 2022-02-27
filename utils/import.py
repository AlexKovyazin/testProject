import re
import os
from settings import ROOT_DIR
from openpyxl import load_workbook
from models import Region, City, User
from pdfminer import high_level


def replace_all(text: str, replace_dict: dict):
    """
    Заменяет в переданной строке символы,
    являющиеся ключами на символы являющиеся значениями.
    :param text: Текст, в котором необходима множественная замена.
    :param replace_dict: Словарь с заменяемыми символами
    (keys - старые значения, values - новые значения).
    :return:
    """
    for old_value, new_value in replace_dict.items():
        text = text.replace(old_value, new_value)
    return text


def import_from_pdf(path_to_file: str):
    """
    Импортирует в БД данные из резюме.
    :param path_to_file: str Путь к файлу.
    :return: None
    """
    text = high_level.extract_text(path_to_file)
    name_pattern = r'([а-яёА-ЯЁ]*)\s([а-яёА-ЯЁ]*)\s([а-яёА-ЯЁ]*)'
    phone_pattern = r'(\+7.[0-9\(\)  ]*)'
    email_pattern = r'(.*@.*)\b'
    city_pattern = r'Проживает: (.*)'

    fullname = re.search(name_pattern, text).groups()
    second_name = fullname[0]
    first_name = fullname[1]
    patronymic = fullname[2]
    raw_phone = re.search(phone_pattern, text).groups()[0]
    email = re.search(email_pattern, text).groups()[0]
    city_name = re.search(city_pattern, text).groups()[0]

    # Преобразовываем данные в формат БД
    replace_dict = {
        '(': '',
        ')': '',
        ' ': '',
        ' ': ''
    }
    clear_phone = replace_all(raw_phone, replace_dict)
    phone = f"{clear_phone[0:2]} " \
            f"{clear_phone[2:5]} " \
            f"{clear_phone[5:8]} " \
            f"{clear_phone[8:10]} " \
            f"{clear_phone[10:12]}"
    try:  # Город может отсутствовать в БД
        city_id = City.get_city_id_by_name(city_name)
        region_id = City.get_region_id(city_id)
    except AttributeError:  # Срабатывает при отсутствии города в БД
        new_city = City(**{'city_name': city_name})
        new_city.create_city()

        city_id = City.get_city_id_by_name(city_name)
        region_id = None  # В резюме не указан регион
        # В связи с этим поле region_id в cities изменено на необязательное

    user_data = {
        'first_name': first_name,
        'second_name': second_name,
        'patronymic': patronymic,
        'phone': phone,
        'email': email,
        'city': city_id,
        'region': region_id
    }
    # Создаём пользователя в БД
    new_user = User(**user_data)
    new_user.create_user()


def import_from_xlsx(path_to_file: str):
    """
    Импортирует в БД данные из .xlsx файла.
    :param path_to_file: str Путь к файлу.
    :return: None
    """
    workbook = load_workbook(path_to_file)
    worksheet = workbook.active

    headers_list = worksheet[worksheet.min_row]
    # Необходим для сопоставления названия колонок
    headers = {i: cell.value for i, cell in enumerate(headers_list)}
    body = worksheet[worksheet.min_row + 1: worksheet.max_row]

    # Собираем данные из таблицы в словари и добавляем в users_data
    users_data = []
    for row in body:
        user_dict = {}
        for i, cell in enumerate(row):
            # Используем headers для правильного распределения значений по столбцам
            user_dict[headers[i]] = cell.value
        region_name = user_dict['region']
        city_name = user_dict['city']

        try:    # Проверяем есть ли регион в БД
            # Заменяем имя на id
            user_dict['region'] = Region.get_region_id_by_name(region_name)
        except AttributeError:  # Добавляем регион в случае отсутствия
            new_region = Region(**{'region_name': user_dict['region'].capitalize()})
            new_region.create_region()
            user_dict['region'] = Region.get_region_id_by_name(region_name)

        try:    # Проверяем есть ли город в БД
            # Заменяем имя на id
            city_id = City.get_city_id_by_name(city_name)
            user_dict['city'] = city_id
            # При наличии города в БД проверяем соответствует ли
            # его регион в импортируемом файле тому, что указан в БД
            if City.get_region_id(city_id) != user_dict['region']:
                print(f"Пользователь {user_dict['second_name']} "
                      f"{user_dict['first_name']} {user_dict['patronymic']} "
                      f"не добавлен в базу данных из-за несоответствия "
                      f"региона указанного города существующим данным")
                # Выполнить необходимые действия...
                continue
        except AttributeError:  # Добавляем город в случае отсутствия
            new_city = City(**{'city_name': user_dict['city'].capitalize(),
                               'region_id': user_dict['region']})
            new_city.create_city()
            user_dict['city'] = City.get_city_id_by_name(city_name)

        users_data.append(user_dict)

    # Создаем пользователей в БД
    for user in users_data:
        new_user = User(**user)
        new_user.create_user()


if __name__ == '__main__':
    # xlsx_path = os.path.join(ROOT_DIR, 'media', 'users_for_import.xlsx')
    # import_from_xlsx(xlsx_path)

    # pdf_path = ''  # Необходимо в переменную записать путь к .pdf файлу импортируемого резюме
    # import_from_pdf(pdf_path)
    pass
