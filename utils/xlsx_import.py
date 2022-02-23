import os
from openpyxl import load_workbook
from models import Region, City, User
from settings import ROOT_DIR


def import_from_xlsx(path_to_file):
    workbook = load_workbook(path_to_file)
    worksheet = workbook.active

    headers_list = worksheet[worksheet.min_row]
    # Необходим для сопоставления названия колонок
    headers = {i: cell.value for i, cell in enumerate(headers_list)}
    body = worksheet[worksheet.min_row + 1: worksheet.max_row]

    # Собираем данные из таблицы в словари и добавляем в users_data
    users_data = []
    for row in body:
        user_dict = {}
        for i, cell in enumerate(row):
            # Используем headers для правильного распределения значений по столбцам
            user_dict[headers[i]] = cell.value
        region_name = user_dict['region']
        city_name = user_dict['city']

        try:    # Проверяем есть ли регион в БД
            # Заменяем имя на id
            user_dict['region'] = Region.get_region_id_by_name(region_name)
        except AttributeError:  # Добавляем регион в случае отсутствия
            new_region = Region(**{'region_name': user_dict['region'].capitalize()})
            new_region.create_region()
            user_dict['region'] = Region.get_region_id_by_name(region_name)

        try:    # Проверяем есть ли город в БД
            # Заменяем имя на id
            city_id = City.get_city_id_by_name(city_name)
            user_dict['city'] = city_id
            # При наличии города в БД проверяем соответствует ли
            # его регион в импортируемом файле тому, что указан в БД
            if City.get_region_id(city_id) != user_dict['region']:
                print(f"Пользователь {user_dict['second_name']} "
                      f"{user_dict['first_name']} {user_dict['patronymic']} "
                      f"не добавлен в базу данных из-за несоответствия "
                      f"региона указанного города существующим данным")
                # Выполнить необходимые действия...
                continue
        except AttributeError:  # Добавляем город в случае отсутствия
            new_city = City(**{'city_name': user_dict['city'].capitalize(),
                               'region_id': user_dict['region']})
            new_city.create_city()
            user_dict['city'] = City.get_city_id_by_name(city_name)

        users_data.append(user_dict)

    # Создаем пользователей в БД
    for user in users_data:
        new_user = User(**user)
        new_user.create_user()


if __name__ == '__main__':
    file = os.path.join(ROOT_DIR, 'media', 'users_for_import.xlsx')
    import_from_xlsx(file)
